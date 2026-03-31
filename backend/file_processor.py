"""Extract plain text from uploaded files for use as Gemini context."""

import io
import logging

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {"pdf", "docx", "doc", "txt", "md", "csv", "html", "htm", "xlsx", "xls", "ods"}
MAX_CHARS = 40_000  # ~10k tokens — keep context manageable


def extract_text(file_bytes: bytes, filename: str) -> str:
    """Return plain text extracted from *file_bytes*.

    Raises ValueError for unsupported types.
    Truncates output to MAX_CHARS with a notice.
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "pdf":
        text = _from_pdf(file_bytes)
    elif ext in ("docx", "doc"):
        text = _from_docx(file_bytes)
    elif ext in ("xlsx", "xls", "ods"):
        text = _from_spreadsheet(file_bytes, ext)
    elif ext in ("txt", "md", "csv", "html", "htm"):
        text = file_bytes.decode("utf-8", errors="replace")
    else:
        raise ValueError(
            f"Tipo de arquivo não suportado: .{ext}. "
            f"Formatos aceitos: {', '.join(sorted(SUPPORTED_EXTENSIONS))}"
        )

    if len(text) > MAX_CHARS:
        text = text[:MAX_CHARS] + "\n\n[... conteúdo truncado por limite de tamanho ...]"

    return text.strip()


def _from_pdf(data: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        raise ValueError("Suporte a PDF indisponível no servidor (pypdf não instalado).")
    reader = PdfReader(io.BytesIO(data))
    pages = []
    for page in reader.pages:
        t = page.extract_text()
        if t:
            pages.append(t)
    return "\n\n".join(pages)


def _from_spreadsheet(data: bytes, ext: str) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("Suporte a planilhas indisponível no servidor (openpyxl não instalado).")

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            line = " | ".join(cells)
            if line.replace("|", "").strip():
                rows.append(line)
        if rows:
            sheets.append(f"=== Planilha: {name} ===\n" + "\n".join(rows))
    return "\n\n".join(sheets) if sheets else "(planilha vazia)"


def _from_docx(data: bytes) -> str:
    try:
        from docx import Document
    except ImportError:
        raise ValueError("Suporte a DOCX indisponível no servidor (python-docx não instalado).")
    doc = Document(io.BytesIO(data))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n\n".join(paragraphs)
